# web_app/services/set_service.py
import logging
import openpyxl
import io
import os
import zipfile
import hashlib
from ..models import db, VocabularySet, User, Flashcard, UserFlashcardProgress
from ..config import FLASHCARD_IMAGES_DIR, FLASHCARD_AUDIO_CACHE_DIR

logger = logging.getLogger(__name__)

class SetService:
    """
    Mô tả: Lớp chứa các hàm xử lý logic nghiệp vụ liên quan đến bộ thẻ (VocabularySet).
    """
    def __init__(self):
        pass

    def get_all_sets_with_details(self):
        """
        Mô tả: Lấy tất cả các bộ thẻ cùng với thông tin chi tiết.
        """
        log_prefix = "[SET_SERVICE|GetAllSets]"
        try:
            sets = db.session.query(
                VocabularySet,
                User.username.label('creator_username'),
                db.func.count(Flashcard.flashcard_id).label('flashcard_count')
            ).outerjoin(User, VocabularySet.creator_user_id == User.user_id)\
             .outerjoin(Flashcard, VocabularySet.set_id == Flashcard.set_id)\
             .group_by(VocabularySet.set_id)\
             .order_by(VocabularySet.title)\
             .all()
            
            results = []
            for set_obj, creator_username, flashcard_count in sets:
                set_obj.creator_username = creator_username or "N/A"
                set_obj.flashcard_count = flashcard_count
                results.append(set_obj)
                
            return results
        except Exception as e:
            logger.error(f"{log_prefix} Lỗi khi truy vấn: {e}", exc_info=True)
            return []

    # BẮT ĐẦU THÊM MỚI: Hàm lấy bộ thẻ theo người tạo
    def get_sets_by_creator_id(self, creator_id):
        """
        Mô tả: Lấy tất cả các bộ thẻ được tạo bởi một người dùng cụ thể.
        """
        log_prefix = f"[SET_SERVICE|GetByCreator|User:{creator_id}]"
        try:
            sets = db.session.query(
                VocabularySet,
                db.func.count(Flashcard.flashcard_id).label('flashcard_count')
            ).filter(VocabularySet.creator_user_id == creator_id)\
             .outerjoin(Flashcard, VocabularySet.set_id == Flashcard.set_id)\
             .group_by(VocabularySet.set_id)\
             .order_by(VocabularySet.title)\
             .all()
            
            results = []
            for set_obj, flashcard_count in sets:
                set_obj.flashcard_count = flashcard_count
                results.append(set_obj)
            return results
        except Exception as e:
            logger.error(f"{log_prefix} Lỗi khi truy vấn: {e}", exc_info=True)
            return []
    # KẾT THÚC THÊM MỚI

    def get_set_by_id(self, set_id):
        """
        Mô tả: Lấy một bộ thẻ cụ thể bằng ID.
        """
        return VocabularySet.query.get(set_id)

    def _process_excel_file(self, vocabulary_set, file_stream, sync_by_id=False):
        """
        Mô tả: Xử lý file Excel được tải lên để thêm hoặc đồng bộ hóa flashcard.
        """
        log_prefix = f"[SET_SERVICE|ProcessExcel|Set:{vocabulary_set.set_id}]"
        
        file_content = file_stream.read()
        in_memory_file = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(in_memory_file)
        
        sheet = workbook.active
        
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
        required_headers = ['front', 'back']
        
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            raise ValueError(f"File Excel thiếu các cột bắt buộc: {', '.join(missing_headers)}.")
        
        column_map = {header: idx for idx, header in enumerate(headers)}
        
        flashcards_from_excel = []

        for row_index, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            row_values = [cell.value for cell in row_cells]
            
            flashcard_id_from_excel = None
            if 'flashcard_id' in column_map and row_values[column_map['flashcard_id']] is not None:
                try:
                    flashcard_id_from_excel = int(row_values[column_map['flashcard_id']])
                except ValueError:
                    logger.warning(f"{log_prefix} Hàng {row_index}: flashcard_id không hợp lệ.")

            front = str(row_values[column_map['front']]).strip() if row_values[column_map['front']] is not None else ''
            back = str(row_values[column_map['back']]).strip() if row_values[column_map['back']] is not None else ''

            if not front:
                logger.warning(f"{log_prefix} Hàng {row_index}: Cột 'front' rỗng.")
            if not back:
                logger.warning(f"{log_prefix} Hàng {row_index}: Cột 'back' rỗng.")

            card_data = {
                'flashcard_id': flashcard_id_from_excel,
                'set_id': vocabulary_set.set_id, 
                'front': front, 
                'back': back
            }
            optional_columns = ['front_audio_content', 'back_audio_content', 'front_img', 'back_img', 'notification_text']
            for col_name in optional_columns:
                if col_name in column_map:
                    col_idx = column_map[col_name]
                    if col_idx < len(row_values):
                        value = row_values[col_idx]
                        card_data[col_name] = str(value).strip() if value is not None else None
            
            flashcards_from_excel.append(card_data)
        
        if sync_by_id:
            existing_flashcards_map = {f.flashcard_id: f for f in vocabulary_set.flashcards}
            excel_flashcard_ids = {f_data['flashcard_id'] for f_data in flashcards_from_excel if f_data['flashcard_id'] is not None}

            for f_data in flashcards_from_excel:
                f_id = f_data.pop('flashcard_id')
                if f_id is not None and f_id in existing_flashcards_map:
                    f_to_update = existing_flashcards_map[f_id]
                    for key, value in f_data.items():
                        setattr(f_to_update, key, value)
                    db.session.add(f_to_update)
                else:
                    new_flashcard = Flashcard(**f_data)
                    db.session.add(new_flashcard)
            
            flashcards_to_delete = [f for f_id, f in existing_flashcards_map.items() if f_id not in excel_flashcard_ids]
            for f in flashcards_to_delete:
                UserFlashcardProgress.query.filter_by(flashcard_id=f.flashcard_id).delete()
                db.session.delete(f)
        else:
            flashcards_to_add = []
            for f_data in flashcards_from_excel:
                f_data.pop('flashcard_id')
                new_flashcard = Flashcard(**f_data)
                flashcards_to_add.append(new_flashcard)
            
            if flashcards_to_add:
                db.session.bulk_save_objects(flashcards_to_add)

    def create_set(self, data, creator_id, file_stream=None):
        """
        Mô tả: Tạo một bộ thẻ mới.
        """
        log_prefix = f"[SET_SERVICE|CreateSet|User:{creator_id}]"
        try:
            new_set = VocabularySet(
                title=data.get('title'),
                description=data.get('description'),
                tags=data.get('tags'),
                is_public=int(data.get('is_public', 1)),
                creator_user_id=creator_id
            )
            db.session.add(new_set)

            if file_stream:
                db.session.flush()
                self._process_excel_file(new_set, file_stream, sync_by_id=False)
            
            db.session.commit()
            return new_set, "success"
        except ValueError as ve:
            db.session.rollback()
            return None, str(ve)
        except Exception as e:
            db.session.rollback()
            if "zip" in str(e).lower():
                 return None, "Lỗi đọc file Excel. Vui lòng đảm bảo file có định dạng .xlsx hợp lệ."
            return None, str(e)

    # BẮT ĐẦU THAY ĐỔI: Thêm user_id và kiểm tra quyền
    def update_set(self, set_id, data, user_id, file_stream=None):
        """
        Mô tả: Cập nhật thông tin và nội dung của một bộ thẻ, có kiểm tra quyền.
        """
        log_prefix = f"[SET_SERVICE|UpdateSet|Set:{set_id}|User:{user_id}]"
        
        set_to_update = self.get_set_by_id(set_id)
        if not set_to_update:
            return None, "set_not_found"
            
        user = User.query.get(user_id)
        if not user:
            return None, "user_not_found"
            
        if user.user_role != 'admin' and set_to_update.creator_user_id != user.user_id:
            logger.warning(f"{log_prefix} Từ chối quyền truy cập.")
            return None, "permission_denied"

        try:
            set_to_update.title = data.get('title', set_to_update.title)
            set_to_update.description = data.get('description', set_to_update.description)
            set_to_update.tags = data.get('tags', set_to_update.tags)
            set_to_update.is_public = int(data.get('is_public', set_to_update.is_public))
            
            if file_stream:
                self._process_excel_file(set_to_update, file_stream, sync_by_id=True)

            db.session.commit()
            return set_to_update, "success"
        except ValueError as ve:
            db.session.rollback()
            return None, str(ve)
        except Exception as e:
            db.session.rollback()
            return None, str(e)
    # KẾT THÚC THAY ĐỔI

    # BẮT ĐẦU THAY ĐỔI: Thêm user_id và kiểm tra quyền
    def delete_set(self, set_id, user_id):
        """
        Mô tả: Xóa một bộ thẻ khỏi cơ sở dữ liệu, có kiểm tra quyền.
        """
        log_prefix = f"[SET_SERVICE|DeleteSet|Set:{set_id}|User:{user_id}]"
        
        set_to_delete = self.get_set_by_id(set_id)
        if not set_to_delete:
            return False, "set_not_found"
            
        user = User.query.get(user_id)
        if not user:
            return False, "user_not_found"
            
        if user.user_role != 'admin' and set_to_delete.creator_user_id != user.user_id:
            logger.warning(f"{log_prefix} Từ chối quyền truy cập.")
            return False, "permission_denied"

        try:
            db.session.delete(set_to_delete)
            db.session.commit()
            return True, "success"
        except Exception as e:
            db.session.rollback()
            return False, str(e)
    # KẾT THÚC THAY ĐỔI

    def export_set_to_excel(self, set_id):
        """
        Mô tả: Xuất tất cả các flashcard của một bộ thẻ ra file Excel.
        """
        set_to_export = self.get_set_by_id(set_id)
        if not set_to_export:
            return None

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = set_to_export.title[:30]

            headers = ['flashcard_id', 'front', 'back', 'front_audio_content', 'back_audio_content', 'front_img', 'back_img', 'notification_text']
            sheet.append(headers)

            for card in set_to_export.flashcards:
                row_data = [
                    card.flashcard_id,
                    card.front,
                    card.back,
                    card.front_audio_content,
                    card.back_audio_content,
                    card.front_img,
                    card.back_img,
                    card.notification_text
                ]
                sheet.append(row_data)
            
            excel_stream = io.BytesIO()
            workbook.save(excel_stream)
            excel_stream.seek(0)
            
            return excel_stream
        except Exception as e:
            logger.error(f"Lỗi khi xuất bộ thẻ ra Excel: {e}", exc_info=True)
            return None

    def export_set_as_zip(self, set_id):
        """
        Mô tả: Xuất một bộ thẻ đầy đủ vào một file ZIP.
        """
        log_prefix = f"[SET_SERVICE|ExportZip|Set:{set_id}]"
        set_to_export = self.get_set_by_id(set_id)
        if not set_to_export:
            return None

        try:
            excel_stream = self.export_set_to_excel(set_id)
            if not excel_stream:
                return None

            zip_stream = io.BytesIO()
            with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr('data.xlsx', excel_stream.getvalue())
                
                added_media_files = set()
                for card in set_to_export.flashcards:
                    for img_attr in ['front_img', 'back_img']:
                        img_filename = getattr(card, img_attr)
                        if img_filename and img_filename not in added_media_files:
                            if not (img_filename.startswith('http://') or img_filename.startswith('https://')):
                                img_path = os.path.join(FLASHCARD_IMAGES_DIR, img_filename)
                                if os.path.exists(img_path):
                                    zf.write(img_path, os.path.join('images', img_filename))
                                    added_media_files.add(img_filename)
                    
                    for audio_attr in ['front_audio_content', 'back_audio_content']:
                        audio_content = getattr(card, audio_attr)
                        if audio_content:
                            content_hash = hashlib.sha1(audio_content.encode('utf-8')).hexdigest()
                            cache_filename = f"{content_hash}.mp3"
                            if cache_filename not in added_media_files:
                                cache_path = os.path.join(FLASHCARD_AUDIO_CACHE_DIR, cache_filename)
                                if os.path.exists(cache_path):
                                    zf.write(cache_path, os.path.join('audio', cache_filename))
                                    added_media_files.add(cache_filename)

            zip_stream.seek(0)
            return zip_stream

        except Exception as e:
            logger.error(f"{log_prefix} Lỗi khi tạo file ZIP: {e}", exc_info=True)
            return None
