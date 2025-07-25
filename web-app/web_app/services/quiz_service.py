# web_app/services/quiz_service.py
import logging
import openpyxl
import io
import random
import time
import hashlib
import os
import zipfile
import requests
from sqlalchemy import func
from ..models import db, QuestionSet, User, QuizQuestion, UserQuizProgress, ScoreLog, QuizPassage
from ..config import (
    SCORE_QUIZ_CORRECT_FIRST_TIME, SCORE_QUIZ_CORRECT_REPEAT,
    QUIZ_MODE_NEW_SEQUENTIAL, QUIZ_MODE_NEW_RANDOM, QUIZ_MODE_REVIEW,
    QUIZ_IMAGES_DIR, QUIZ_AUDIO_CACHE_DIR
)

logger = logging.getLogger(__name__)

def _sort_sets_by_progress(set_items, total_key, completed_key):
    """
    Mô tả: Sắp xếp danh sách các bộ dựa trên tiến độ hoàn thành.
    """
    def custom_sort_key(set_item):
        total = getattr(set_item, total_key, 0)
        completed = getattr(set_item, completed_key, 0)
        title = getattr(set_item, 'title', '')

        if total == 0:
            return (float('-inf'), title) 

        percentage = (completed * 100 / total)

        if percentage == 100:
            return (0, title) 
        
        return (-percentage, title)

    return sorted(set_items, key=custom_sort_key)

class QuizService:
    """
    Mô tả: Lớp chứa các hàm xử lý logic nghiệp vụ liên quan đến bộ câu hỏi trắc nghiệm.
    """
    def __init__(self):
        pass

    def get_categorized_question_sets_for_user(self, user_id):
        """
        Mô tả: Lấy và phân loại các bộ câu hỏi thành "đã bắt đầu" và "mới".
        """
        try:
            started_set_ids_query = db.session.query(QuizQuestion.set_id).join(
                UserQuizProgress, UserQuizProgress.question_id == QuizQuestion.question_id
            ).filter(UserQuizProgress.user_id == user_id).distinct()
            
            started_set_ids = {row[0] for row in started_set_ids_query.all()}
            
            started_sets_raw = []
            if started_set_ids:
                total_questions_map = dict(db.session.query(
                    QuizQuestion.set_id, func.count(QuizQuestion.question_id)
                ).filter(QuizQuestion.set_id.in_(started_set_ids)).group_by(QuizQuestion.set_id).all())
                
                answered_questions_map = dict(db.session.query(
                    QuizQuestion.set_id, func.count(UserQuizProgress.progress_id)
                ).join(QuizQuestion).filter(
                    UserQuizProgress.user_id == user_id,
                    QuizQuestion.set_id.in_(started_set_ids)
                ).group_by(QuizQuestion.set_id).all())

                set_objects = QuestionSet.query.filter(QuestionSet.set_id.in_(started_set_ids)).all()
                for s in set_objects:
                    s.total_questions = total_questions_map.get(s.set_id, 0)
                    s.answered_questions = answered_questions_map.get(s.set_id, 0)
                    s.creator_username = s.creator.username if s.creator else "N/A"
                    started_sets_raw.append(s)

            started_sets = _sort_sets_by_progress(started_sets_raw, 
                                                total_key='total_questions', 
                                                completed_key='answered_questions')
            
            new_sets_query = QuestionSet.query.filter(
                QuestionSet.is_public == True,
                ~QuestionSet.set_id.in_(started_set_ids)
            ).order_by(QuestionSet.title.asc())
            new_sets = new_sets_query.all()
            for s in new_sets:
                 s.creator_username = s.creator.username if s.creator else "N/A"

            return started_sets, new_sets

        except Exception as e:
            logger.error(f"Lỗi khi phân loại bộ câu hỏi: {e}", exc_info=True)
            return [], []

    def get_next_question_group_for_user(self, user_id, set_id, mode):
        """
        Mô tả: Lấy nhóm câu hỏi tiếp theo cho người dùng.
        """
        all_q_ids_in_set = {row[0] for row in QuizQuestion.query.with_entities(QuizQuestion.question_id).filter_by(set_id=set_id).all()}
        if not all_q_ids_in_set:
            return None, None

        answered_q_ids = {row[0] for row in UserQuizProgress.query.with_entities(UserQuizProgress.question_id).filter_by(user_id=user_id).join(QuizQuestion).filter(QuizQuestion.set_id == set_id).all()}

        user = User.query.get(user_id)
        if user and user.current_question_set_id == set_id:
            last_answered_progress = UserQuizProgress.query.filter_by(user_id=user_id)\
                                     .join(QuizQuestion).filter(QuizQuestion.set_id == set_id)\
                                     .order_by(UserQuizProgress.last_answered.desc())\
                                     .first()
            
            if last_answered_progress and last_answered_progress.question.passage_id:
                current_passage_id = last_answered_progress.question.passage_id
                
                unanswered_in_passage_count = QuizQuestion.query.filter(
                    QuizQuestion.set_id == set_id,
                    QuizQuestion.passage_id == current_passage_id,
                    ~QuizQuestion.progresses.any(user_id=user_id)
                ).count()

                if unanswered_in_passage_count > 0:
                    questions_in_passage = QuizQuestion.query.filter(
                        QuizQuestion.set_id == set_id,
                        QuizQuestion.passage_id == current_passage_id
                    ).order_by(QuizQuestion.passage_order.asc(), QuizQuestion.question_id.asc()).all()
                    
                    return questions_in_passage, questions_in_passage[0].passage
        
        next_question_id = None
        
        if mode == QUIZ_MODE_NEW_SEQUENTIAL:
            new_q_ids = all_q_ids_in_set - answered_q_ids
            if new_q_ids:
                next_question_id = min(new_q_ids)

        elif mode == QUIZ_MODE_NEW_RANDOM:
            new_q_ids = all_q_ids_in_set - answered_q_ids
            if new_q_ids:
                next_question_id = random.choice(list(new_q_ids))
        
        elif mode == QUIZ_MODE_REVIEW:
            if answered_q_ids:
                next_question_id = random.choice(list(answered_q_ids))
        
        if next_question_id:
            next_question = self.get_question_by_id(next_question_id)
            if next_question.passage_id:
                questions_in_passage = QuizQuestion.query.filter(
                    QuizQuestion.set_id == set_id,
                    QuizQuestion.passage_id == next_question.passage_id
                ).order_by(QuizQuestion.passage_order.asc(), QuizQuestion.question_id.asc()).all()
                return questions_in_passage, next_question.passage
            else:
                return [next_question], None
        
        return None, None

    def process_user_answers(self, user_id, answers_data):
        """
        Mô tả: Xử lý nhiều câu trả lời của người dùng.
        """
        results = []
        try:
            for answer in answers_data:
                question_id = answer.get('question_id')
                selected_option = answer.get('selected_option')

                question = self.get_question_by_id(question_id)
                if not question:
                    results.append({'question_id': question_id, 'status': 'error'})
                    continue

                is_correct = (selected_option == question.correct_answer)
                
                progress = UserQuizProgress.query.filter_by(user_id=user_id, question_id=question_id).first()
                if not progress:
                    progress = UserQuizProgress(user_id=user_id, question_id=question_id)
                    db.session.add(progress)
                
                if progress.times_correct is None: progress.times_correct = 0
                if progress.times_incorrect is None: progress.times_incorrect = 0
                if progress.correct_streak is None: progress.correct_streak = 0

                progress.last_answered = int(time.time())
                score_change = 0
                
                is_first_correct = False
                if is_correct:
                    is_first_correct = (progress.times_correct == 0)
                    score_change = SCORE_QUIZ_CORRECT_FIRST_TIME if is_first_correct else SCORE_QUIZ_CORRECT_REPEAT
                    progress.times_correct += 1
                    progress.correct_streak += 1
                else:
                    progress.times_incorrect += 1
                    progress.correct_streak = 0
                
                if progress.correct_streak >= 3:
                    progress.is_mastered = True
                else:
                    progress.is_mastered = False

                if score_change > 0:
                    user = User.query.get(user_id)
                    user.score = (user.score or 0) + score_change
                    db.session.add(user)
                    
                    reason = f"quiz_answer_{'first_correct' if is_first_correct else 'correct'}"
                    score_log = ScoreLog(user_id=user_id, score_change=score_change, timestamp=int(time.time()), reason=reason, source_type='quiz')
                    db.session.add(score_log)
                
                db.session.add(progress)
                
                results.append({'question_id': question_id, 'is_correct': is_correct, 'correct_answer': question.correct_answer, 'guidance': question.guidance or '', 'status': 'success'})
            
            db.session.commit()
            return results
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Lỗi khi xử lý câu trả lời: {e}", exc_info=True)
            return [{'status': 'error', 'message': 'Lỗi server nội bộ.'}]

    def get_all_question_sets_with_details(self):
        """
        Mô tả: Lấy tất cả các bộ câu hỏi cùng với thông tin chi tiết.
        """
        try:
            sets = db.session.query(
                QuestionSet,
                User.username.label('creator_username'),
                db.func.count(QuizQuestion.question_id).label('question_count')
            ).outerjoin(User, QuestionSet.creator_user_id == User.user_id)\
             .outerjoin(QuizQuestion, QuestionSet.set_id == QuizQuestion.set_id)\
             .group_by(QuestionSet.set_id).order_by(QuestionSet.title).all()
            
            results = []
            for set_obj, creator_username, question_count in sets:
                set_obj.creator_username = creator_username or "N/A"
                set_obj.question_count = question_count
                results.append(set_obj)
            return results
        except Exception as e:
            logger.error(f"Lỗi khi truy vấn: {e}", exc_info=True)
            return []

    # BẮT ĐẦU THÊM MỚI: Hàm lấy bộ câu hỏi theo người tạo
    def get_question_sets_by_creator_id(self, creator_id):
        """
        Mô tả: Lấy tất cả các bộ câu hỏi được tạo bởi một người dùng cụ thể.
        """
        log_prefix = f"[QUIZ_SERVICE|GetByCreator|User:{creator_id}]"
        try:
            sets = db.session.query(
                QuestionSet,
                db.func.count(QuizQuestion.question_id).label('question_count')
            ).filter(QuestionSet.creator_user_id == creator_id)\
             .outerjoin(QuizQuestion, QuestionSet.set_id == QuizQuestion.set_id)\
             .group_by(QuestionSet.set_id)\
             .order_by(QuestionSet.title)\
             .all()
            
            results = []
            for set_obj, question_count in sets:
                set_obj.question_count = question_count
                results.append(set_obj)
            return results
        except Exception as e:
            logger.error(f"{log_prefix} Lỗi khi truy vấn: {e}", exc_info=True)
            return []
    # KẾT THÚC THÊM MỚI

    def get_question_set_by_id(self, set_id):
        return QuestionSet.query.get(set_id)

    def get_question_by_id(self, question_id):
        return QuizQuestion.query.get(question_id)

    def update_question(self, question_id, data, user_id):
        """
        Mô tả: Cập nhật nội dung của một câu hỏi.
        """
        question = self.get_question_by_id(question_id)
        if not question:
            return None, "question_not_found"

        user = User.query.get(user_id)
        if not user:
            return None, "user_not_found"

        set_creator_id = question.question_set.creator_user_id
        if user.user_role != 'admin' and user.user_id != set_creator_id:
            return None, "permission_denied"

        try:
            question.pre_question_text = data.get('pre_question_text') or None
            question.question = data.get('question') or None
            question.question_image_file = data.get('question_image_file') or None
            question.question_audio_file = data.get('question_audio_file') or None
            question.option_a = data.get('option_a')
            question.option_b = data.get('option_b')
            question.option_c = data.get('option_c') or None
            question.option_d = data.get('option_d') or None
            question.correct_answer = data.get('correct_answer', question.correct_answer).upper()
            question.guidance = data.get('guidance') or None
            
            if 'passage_order' in data and data['passage_order'] is not None:
                try:
                    question.passage_order = int(data['passage_order'])
                except (ValueError, TypeError):
                    question.passage_order = None
            else:
                question.passage_order = None
            
            db.session.commit()
            return question, "success"
        except Exception as e:
            db.session.rollback()
            return None, str(e)

    def _process_excel_file(self, question_set, file_stream, sync=False):
        """
        Mô tả: Xử lý file Excel để thêm hoặc đồng bộ hóa câu hỏi.
        """
        file_content = file_stream.read()
        in_memory_file = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(in_memory_file)
        sheet = workbook.active
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
        required_headers = ['question', 'option_a', 'option_b', 'correct_answer_text']
        
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            raise ValueError(f"File Excel thiếu các cột bắt buộc: {', '.join(missing_headers)}.")
        
        column_map = {header: idx for idx, header in enumerate(headers)}
        questions_from_excel = []
        passage_content_to_id_map = {p.passage_hash: p.passage_id for p in QuizPassage.query.all()}

        for row_index, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            row_values = [cell.value for cell in row_cells]
            question_id_from_excel = None
            if 'question_id' in column_map and row_values[column_map['question_id']] is not None:
                try:
                    question_id_from_excel = int(row_values[column_map['question_id']])
                except ValueError:
                    pass

            question_text = str(row_values[column_map.get('question')]).strip() if column_map.get('question') is not None and row_values[column_map.get('question')] is not None else ''
            question_image_file = str(row_values[column_map.get('question_image_file')]).strip() if column_map.get('question_image_file') is not None and row_values[column_map.get('question_image_file')] is not None else ''
            question_audio_file = str(row_values[column_map.get('question_audio_file')]).strip() if column_map.get('question_audio_file') is not None and row_values[column_map.get('question_audio_file')] is not None else ''

            if not question_text and not question_image_file and not question_audio_file:
                continue

            option_a_text = str(row_values[column_map['option_a']]).strip()
            option_b_text = str(row_values[column_map['option_b']]).strip()
            option_c_text = str(row_values[column_map.get('option_c')]).strip() if column_map.get('option_c') is not None and row_values[column_map.get('option_c')] is not None else ''
            option_d_text = str(row_values[column_map.get('option_d')]).strip() if column_map.get('option_d') is not None and row_values[column_map.get('option_d')] is not None else ''
            correct_answer_text = str(row_values[column_map['correct_answer_text']]).strip()

            determined_answer = None
            if correct_answer_text and option_a_text and correct_answer_text == option_a_text: determined_answer = 'A'
            elif correct_answer_text and option_b_text and correct_answer_text == option_b_text: determined_answer = 'B'
            elif correct_answer_text and option_c_text and correct_answer_text == option_c_text: determined_answer = 'C'
            elif correct_answer_text and option_d_text and correct_answer_text == option_d_text: determined_answer = 'D'

            if not determined_answer: continue

            passage_text_from_excel = str(row_values[column_map.get('passage_text')]).strip() if column_map.get('passage_text') is not None and row_values[column_map.get('passage_text')] is not None else ''
            passage_order_from_excel = None
            if 'passage_order' in column_map and row_values[column_map['passage_order']] is not None:
                try:
                    passage_order_from_excel = int(row_values[column_map['passage_order']])
                except (ValueError, TypeError):
                    pass

            passage_id_for_db = None
            if passage_text_from_excel:
                passage_hash = hashlib.sha256(passage_text_from_excel.encode('utf-8')).hexdigest()
                if passage_hash not in passage_content_to_id_map:
                    new_passage = QuizPassage(passage_content=passage_text_from_excel, passage_hash=passage_hash)
                    db.session.add(new_passage)
                    db.session.flush()
                    passage_content_to_id_map[passage_hash] = new_passage.passage_id
                passage_id_for_db = passage_content_to_id_map[passage_hash]
            
            question_data = {
                'question_id': question_id_from_excel, 'question': question_text or None,
                'option_a': option_a_text, 'option_b': option_b_text, 'option_c': option_c_text or None,
                'option_d': option_d_text or None, 'correct_answer': determined_answer, 
                'pre_question_text': str(row_values[column_map.get('pre_question_text')]).strip() if column_map.get('pre_question_text') is not None and row_values[column_map.get('pre_question_text')] is not None else None,
                'guidance': str(row_values[column_map.get('guidance')]).strip() if column_map.get('guidance') is not None and row_values[column_map.get('guidance')] is not None else None,
                'question_image_file': question_image_file or None, 'question_audio_file': question_audio_file or None,
                'passage_id': passage_id_for_db, 'passage_order': passage_order_from_excel
            }
            questions_from_excel.append(question_data)

        if sync:
            existing_questions_map = {q.question_id: q for q in question_set.questions}
            excel_question_ids = {q_data['question_id'] for q_data in questions_from_excel if q_data['question_id'] is not None}

            for q_data in questions_from_excel:
                q_id = q_data.pop('question_id')
                if q_id is not None and q_id in existing_questions_map:
                    q_to_update = existing_questions_map[q_id]
                    for key, value in q_data.items():
                        setattr(q_to_update, key, value)
                    db.session.add(q_to_update)
                else:
                    new_question = QuizQuestion(set_id=question_set.set_id, **q_data)
                    db.session.add(new_question)
            
            questions_to_delete = [q for q_id, q in existing_questions_map.items() if q_id not in excel_question_ids]
            for q in questions_to_delete:
                db.session.delete(q)
        else:
            questions_to_add = []
            for q_data in questions_from_excel:
                q_data.pop('question_id')
                new_question = QuizQuestion(set_id=question_set.set_id, **q_data)
                questions_to_add.append(new_question)
            
            if questions_to_add:
                db.session.bulk_save_objects(questions_to_add)

    def create_question_set(self, data, creator_id, file_stream=None):
        """
        Mô tả: Tạo một bộ câu hỏi mới.
        """
        try:
            new_set = QuestionSet(
                title=data.get('title'), description=data.get('description'),
                is_public=int(data.get('is_public', 1)), creator_user_id=creator_id
            )
            db.session.add(new_set)
            if file_stream:
                db.session.flush()
                self._process_excel_file(new_set, file_stream, sync=False)
            db.session.commit()
            return new_set, "success"
        except ValueError as ve:
            db.session.rollback()
            return None, str(ve)
        except Exception as e:
            db.session.rollback()
            if "zip" in str(e).lower():
                 return None, "Lỗi đọc file Excel."
            return None, str(e)

    # BẮT ĐẦU THAY ĐỔI: Thêm user_id và kiểm tra quyền
    def update_question_set(self, set_id, data, user_id, file_stream=None):
        """
        Mô tả: Cập nhật thông tin và nội dung của một bộ câu hỏi, có kiểm tra quyền.
        """
        log_prefix = f"[QUIZ_SERVICE|UpdateSet|Set:{set_id}|User:{user_id}]"
        set_to_update = self.get_question_set_by_id(set_id)
        if not set_to_update:
            return None, "set_not_found"
            
        user = User.query.get(user_id)
        if not user:
            return None, "user_not_found"
            
        if user.user_role != 'admin' and set_to_update.creator_user_id != user.user_id:
            logger.warning(f"{log_prefix} Từ chối quyền truy cập.")
            return None, "permission_denied"

        try:
            set_to_update.title = data.get('title', set_to_update.title)
            set_to_update.description = data.get('description', set_to_update.description)
            set_to_update.is_public = int(data.get('is_public', set_to_update.is_public))
            
            if file_stream:
                self._process_excel_file(set_to_update, file_stream, sync=True)

            db.session.commit()
            return set_to_update, "success"
        except ValueError as ve:
            db.session.rollback()
            return None, str(ve)
        except Exception as e:
            db.session.rollback()
            return None, str(e)
    # KẾT THÚC THAY ĐỔI

    # BẮT ĐẦU THAY ĐỔI: Thêm user_id và kiểm tra quyền
    def delete_question_set(self, set_id, user_id):
        """
        Mô tả: Xóa một bộ câu hỏi, có kiểm tra quyền.
        """
        log_prefix = f"[QUIZ_SERVICE|DeleteSet|Set:{set_id}|User:{user_id}]"
        set_to_delete = self.get_question_set_by_id(set_id)
        if not set_to_delete:
            return False, "set_not_found"
            
        user = User.query.get(user_id)
        if not user:
            return False, "user_not_found"
            
        if user.user_role != 'admin' and set_to_delete.creator_user_id != user.user_id:
            logger.warning(f"{log_prefix} Từ chối quyền truy cập.")
            return False, "permission_denied"

        try:
            db.session.delete(set_to_delete)
            db.session.commit()
            return True, "success"
        except Exception as e:
            db.session.rollback()
            return False, str(e)
    # KẾT THÚC THAY ĐỔI

    def export_set_to_excel(self, set_id):
        """
        Mô tả: Xuất một bộ câu hỏi ra file Excel.
        """
        set_to_export = self.get_question_set_by_id(set_id)
        if not set_to_export:
            return None

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            invalid_chars = r'\/?*[]:'
            sanitized_title = "".join(c for c in set_to_export.title if c not in invalid_chars)
            sheet.title = sanitized_title[:30]

            headers = ['question_id', 'pre_question_text', 'question', 'option_a', 'option_b', 'option_c', 'option_d', 
                       'correct_answer_text', 'guidance', 'question_image_file', 'question_audio_file',
                       'passage_text', 'passage_order']
            sheet.append(headers)

            sorted_questions = sorted(set_to_export.questions, 
                                      key=lambda q: (q.passage_id if q.passage_id is not None else float('inf'), 
                                                     q.passage_order if q.passage_order is not None else float('inf'), 
                                                     q.question_id))

            for q in sorted_questions:
                correct_answer_text = ""
                if q.correct_answer == 'A': correct_answer_text = q.option_a
                elif q.correct_answer == 'B': correct_answer_text = q.option_b
                elif q.correct_answer == 'C': correct_answer_text = q.option_c
                elif q.correct_answer == 'D': correct_answer_text = q.option_d

                passage_content_to_export = q.passage.passage_content if q.passage else None

                row_data = [
                    q.question_id, q.pre_question_text, q.question,
                    q.option_a, q.option_b, q.option_c, q.option_d,
                    correct_answer_text, q.guidance,
                    q.question_image_file, q.question_audio_file,
                    passage_content_to_export, q.passage_order
                ]
                sheet.append(row_data)
            
            excel_stream = io.BytesIO()
            workbook.save(excel_stream)
            excel_stream.seek(0)
            
            return excel_stream
        except Exception as e:
            logger.error(f"Lỗi khi xuất bộ câu hỏi ra Excel: {e}", exc_info=True)
            return None

    def get_quiz_set_stats_for_user(self, user_id, set_id):
        """
        Mô tả: Lấy các số liệu thống kê chi tiết của một bộ câu hỏi quiz.
        """
        stats = {'set_id': set_id, 'set_title': 'N/A'}
        question_set = QuestionSet.query.get(set_id)
        if not question_set:
            return stats

        stats['set_title'] = question_set.title
        stats['total_questions'] = QuizQuestion.query.filter_by(set_id=set_id).count()
        progress_in_quiz_set = UserQuizProgress.query.join(QuizQuestion).filter(QuizQuestion.set_id == set_id, UserQuizProgress.user_id == user_id)
        stats['answered_questions'] = progress_in_quiz_set.count()
        stats['correct_answers'] = progress_in_quiz_set.filter(UserQuizProgress.times_correct > 0).count()
        stats['incorrect_answers'] = progress_in_quiz_set.filter(UserQuizProgress.times_incorrect > 0).count()
        stats['mastered_questions'] = progress_in_quiz_set.filter(UserQuizProgress.is_mastered == True).count()
        stats['unanswered_questions'] = stats['total_questions'] - stats['answered_questions']
        return stats

    def export_question_set_as_zip(self, set_id):
        """
        Mô tả: Xuất một bộ câu hỏi đầy đủ vào một file ZIP.
        """
        log_prefix = f"[QUIZ_SERVICE|ExportZip|Set:{set_id}]"
        set_to_export = self.get_question_set_by_id(set_id)
        if not set_to_export:
            return None

        try:
            excel_stream = self.export_set_to_excel(set_id)
            if not excel_stream:
                return None

            zip_stream = io.BytesIO()
            with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr('data.xlsx', excel_stream.getvalue())

                added_media_files = set()
                for question in set_to_export.questions:
                    img_filename = question.question_image_file
                    if img_filename and img_filename not in added_media_files:
                        if img_filename.startswith(('http://', 'https://')):
                            try:
                                response = requests.get(img_filename, stream=True, timeout=10)
                                if response.status_code == 200:
                                    safe_filename = hashlib.sha1(img_filename.encode()).hexdigest() + os.path.splitext(img_filename)[1]
                                    zf.writestr(os.path.join('images', safe_filename), response.content)
                                    added_media_files.add(img_filename)
                            except requests.exceptions.RequestException:
                                pass
                        else:
                            img_path = os.path.join(QUIZ_IMAGES_DIR, img_filename)
                            if os.path.exists(img_path):
                                zf.write(img_path, os.path.join('images', img_filename))
                                added_media_files.add(img_filename)

                    audio_filename = question.question_audio_file
                    if audio_filename and audio_filename not in added_media_files:
                        if not audio_filename.startswith(('http://', 'https://')):
                            audio_path = os.path.join(QUIZ_AUDIO_CACHE_DIR, audio_filename)
                            if os.path.exists(audio_path):
                                zf.write(audio_path, os.path.join('audio', audio_filename))
                                added_media_files.add(audio_filename)
            
            zip_stream.seek(0)
            return zip_stream

        except Exception as e:
            logger.error(f"{log_prefix} Lỗi khi tạo file ZIP: {e}", exc_info=True)
            return None
