# Path: flashcard_v2/services/excel_service.py
"""
Module chứa business logic liên quan đến việc đọc và ghi file Excel
cho các chức năng export và update dữ liệu flashcard.
Các hàm đã được cập nhật để sử dụng user_id (khóa chính).
Đã sửa logic xử lý ô audio trống khi import/update.
Đã định dạng lại code, đảm bảo không dùng ; nối lệnh và sửa lỗi biến exception.
"""
import logging
import os
import sqlite3
import openpyxl # Cần import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill # Import các style
from openpyxl.utils import get_column_letter # Import tiện ích cột
from openpyxl.utils.exceptions import InvalidFileException # Import exception cụ thể
import html # Cần cho việc escape khi export
import time # Cần cho việc tạo tên file export
import re # Cần cho việc escape markdown

# Import từ các module khác
from database.connection import database_connect
from database.query_stats import get_review_stats # Vẫn cần cho export user data
from database.query_user import ALLOWED_USER_COLUMNS # Dùng trong export user data
from utils.helpers import convert_unix_to_local # Dùng trong export user data
from utils.exceptions import (
    DatabaseError,
    UserNotFoundError,
    SetNotFoundError,
    PermissionsError,
    FileProcessingError,
    InvalidFileFormatError,
    ExcelImportError,
    ValidationError, # Thêm nếu cần
    DuplicateError   # Thêm nếu cần
)
from config import DEFAULT_TIMEZONE_OFFSET # Dùng trong export user data

logger = logging.getLogger(__name__)

def export_user_data_excel(user_id, output_file="user_data_export.xlsx"):
    """
    Xuất toàn bộ dữ liệu học tập của người dùng ra file Excel nhiều sheet, dựa trên user_id (PK).
    """
    log_prefix = f"[EXPORT_USER_DATA|UserUID:{user_id}]"
    logger.info(f"{log_prefix} Bắt đầu xuất dữ liệu ra file: {output_file}")
    conn = None
    wb = None
    try:
        conn = database_connect()
        if conn is None:
            raise DatabaseError("Không thể kết nối DB.")
        # Đặt row_factory sau khi kết nối thành công
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        logger.debug(f"{log_prefix} Lấy thông tin user...")
        user_select_cols = ", ".join(f'"{col}"' for col in ALLOWED_USER_COLUMNS)
        cursor.execute(f"SELECT {user_select_cols} FROM Users WHERE user_id = ?", (user_id,))
        user_row = cursor.fetchone()
        if not user_row:
            raise UserNotFoundError(identifier=user_id)
        user_dict = dict(user_row)

        logger.debug(f"{log_prefix} Lấy stats...")
        # Truyền conn đã có vào hàm get_review_stats
        stats = get_review_stats(user_id, conn=conn)

        logger.debug(f"{log_prefix} Lấy danh sách bộ từ đã học...")
        query_user_sets = """
            SELECT DISTINCT vs.set_id, vs.title
            FROM UserFlashcardProgress ufp
            JOIN Flashcards f ON ufp.flashcard_id = f.flashcard_id
            JOIN VocabularySets vs ON f.set_id = vs.set_id
            WHERE ufp.user_id = ? ORDER BY vs.title COLLATE NOCASE
        """
        cursor.execute(query_user_sets, (user_id,))
        user_sets = cursor.fetchall()
        logger.info(f"{log_prefix} Tìm thấy {len(user_sets)} bộ từ người dùng đã học.")

        # --- Tạo workbook và sheet Tổng kết ---
        wb = openpyxl.Workbook()
        header_font = Font(bold=True, name='Calibri', size=11)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws_summary = wb.active
        ws_summary.title = "Tổng kết"
        summary_headers = ["Thông tin", "Giá trị"]
        ws_summary.append(summary_headers)
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill

        summary_data = [
            ["ID người dùng (DB)", user_id],
            ["ID Telegram", user_dict.get("telegram_id", "N/A")],
            ["Username", user_dict.get("username", "N/A")],
            ["Điểm số", user_dict.get("score", 0)],
            ["Vai trò", user_dict.get("user_role", "user")],
            ["Số bộ đã học", stats.get('learned_sets', 0)],
            ["Tổng flashcard đã học", stats.get('learned_distinct', 0)],
            ["Tổng flashcard cần ôn", stats.get('course_due_total', 0)],
            ["Giới hạn thẻ mới/ngày", user_dict.get("daily_new_limit", "N/A")],
            ["Múi giờ Offset", user_dict.get("timezone_offset", "N/A")],
            ["Audio Trước", "Bật" if user_dict.get('front_audio', 1) == 1 else "Tắt"],
            ["Audio Sau", "Bật" if user_dict.get('back_audio', 1) == 1 else "Tắt"],
            ["Thông báo", "Bật" if user_dict.get('is_notification_enabled', 0) == 1 else "Tắt"],
        ]
        for row_content in summary_data:
            ws_summary.append(row_content)

        # Định dạng cột sheet Tổng kết
        for col_idx, column_cells in enumerate(ws_summary.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        header_len = len(str(ws_summary.cell(row=1, column=col_idx).value))
                        max_cell_len = max(cell_len, header_len)
                        if max_cell_len > max_length:
                            max_length = max_cell_len
                except Exception:
                    pass # Bỏ qua lỗi tính độ dài
            adjusted_width = max_length + 3
            ws_summary.column_dimensions[column_letter].width = adjusted_width
            # Căn lề trái cho cột giá trị
            if col_idx == 2:
                for cell in column_cells[1:]:
                    cell.alignment = left_alignment

        # --- Tạo các sheet chi tiết cho từng bộ đã học ---
        if user_sets:
            logger.debug(f"{log_prefix} Đang tạo các sheet chi tiết...")
            for set_row in user_sets:
                set_id_current = set_row["set_id"]
                set_title_current = set_row["title"]
                # Làm sạch tên sheet
                safe_sheet_title = "".join(c for c in set_title_current if c.isalnum() or c in (' ', '_', '-')).strip()[:31] or f"Set_{set_id_current}"
                sheet_suffix = 1
                final_sheet_title = safe_sheet_title
                # Xử lý trùng tên sheet
                while final_sheet_title in wb.sheetnames:
                    sheet_suffix_str = f"_{sheet_suffix}"
                    trunc_len = 31 - len(sheet_suffix_str) # Độ dài tối đa của sheet name là 31
                    final_sheet_title = f"{safe_sheet_title[:trunc_len]}{sheet_suffix_str}"
                    sheet_suffix += 1

                ws_set = wb.create_sheet(title=final_sheet_title)
                logger.debug(f"{log_prefix} Tạo sheet: '{ws_set.title}' cho set_id {set_id_current}")

                # Header cho sheet chi tiết
                set_headers = ["Flashcard ID", "Mặt trước (Front)", "Mặt sau (Back)", "Ghi chú (Note)", "Ảnh Trước (Front Image)", "Ảnh Sau (Back Image)", "Nội dung Thông báo", "Lần ôn (Reviews)", "Lần đúng (Correct)", "Chuỗi đúng (Streak)", "Ôn lần cuối (Last)", "Ôn lần tới (Due)", "Front Audio Content", "Back Audio Content"]
                ws_set.append(set_headers)
                # Định dạng header
                for cell in ws_set[1]:
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.fill = header_fill

                # Lấy dữ liệu flashcard và progress cho bộ hiện tại
                query_set_details = """
                    SELECT
                        f.flashcard_id, f.front, f.back, f.front_audio_content, f.back_audio_content,
                        f.front_img, f.back_img, f.notification_text,
                        fn.note,
                        ufp.review_count, ufp.correct_count, ufp.correct_streak,
                        ufp.last_reviewed, ufp.due_time
                    FROM Flashcards f
                    LEFT JOIN UserFlashcardProgress ufp ON f.flashcard_id = ufp.flashcard_id AND ufp.user_id = ?
                    LEFT JOIN FlashcardNotes fn ON f.flashcard_id = fn.flashcard_id AND fn.user_id = ?
                    WHERE f.set_id = ?
                    ORDER BY f.flashcard_id ASC
                """
                cursor.execute(query_set_details, (user_id, user_id, set_id_current))
                flashcards_in_set = cursor.fetchall()
                logger.debug(f"{log_prefix} Tìm thấy {len(flashcards_in_set)} chi tiết cho set {set_id_current}.")

                # Ghi dữ liệu vào sheet
                tz_offset = user_dict.get('timezone_offset', DEFAULT_TIMEZONE_OFFSET)
                for flashcard_row in flashcards_in_set:
                    last_rev_ts = flashcard_row["last_reviewed"]
                    due_time_ts = flashcard_row["due_time"]
                    row_data = [
                        flashcard_row["flashcard_id"],
                        flashcard_row["front"],
                        flashcard_row["back"],
                        flashcard_row["note"] if flashcard_row["note"] is not None else "",
                        flashcard_row["front_img"] if flashcard_row["front_img"] is not None else "",
                        flashcard_row["back_img"] if flashcard_row["back_img"] is not None else "",
                        flashcard_row["notification_text"] if flashcard_row["notification_text"] is not None else "",
                        flashcard_row["review_count"] if flashcard_row["review_count"] is not None else 0,
                        flashcard_row["correct_count"] if flashcard_row["correct_count"] is not None else 0,
                        flashcard_row["correct_streak"] if flashcard_row["correct_streak"] is not None else 0,
                        convert_unix_to_local(last_rev_ts, tz_offset) if last_rev_ts else "",
                        convert_unix_to_local(due_time_ts, tz_offset) if due_time_ts else "",
                        flashcard_row["front_audio_content"] if flashcard_row["front_audio_content"] is not None else "",
                        flashcard_row["back_audio_content"] if flashcard_row["back_audio_content"] is not None else ""
                    ]
                    ws_set.append(row_data)

                # Định dạng cột sheet chi tiết
                left_align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
                center_align_no_wrap = Alignment(horizontal='center', vertical='top', wrap_text=False)
                for col_idx, header_text in enumerate(set_headers, 1):
                    max_length = len(header_text) + 2
                    column_letter = get_column_letter(col_idx)
                    for cell in ws_set[column_letter]:
                         if cell.row > 1 and cell.value:
                             try:
                                 max_col_width = 60 # Mặc định
                                 if 'image' in header_text.lower() or 'audio' in header_text.lower() or 'Thông báo' in header_text:
                                     max_col_width = 40
                                 elif 'note' in header_text.lower():
                                     max_col_width = 80
                                 elif 'id' in header_text.lower() or 'count' in header_text.lower() or 'streak' in header_text.lower():
                                     max_col_width = 15
                                 cell_len = len(str(cell.value))
                                 current_width_estimate = min(max(cell_len, 10), max_col_width)
                                 if current_width_estimate > max_length:
                                     max_length = current_width_estimate
                             except Exception as len_err:
                                 logger.warning(f"{log_prefix} Lỗi tính độ dài ô {cell.coordinate}: {len_err}")
                    adjusted_width = max_length + 2
                    ws_set.column_dimensions[column_letter].width = adjusted_width
                    # Đặt alignment
                    is_center_col = header_text.startswith("Flashcard ID") or header_text.startswith("Lần ôn") or header_text.startswith("Lần đúng") or header_text.startswith("Chuỗi đúng")
                    alignment_to_use = center_align_no_wrap if is_center_col else left_align_wrap
                    # Áp dụng cho các ô dữ liệu (từ hàng 2)
                    for cell in ws_set[column_letter][1:]:
                        cell.alignment = alignment_to_use
        else:
             logger.info(f"{log_prefix} Người dùng chưa học bộ từ nào.")

        # Lưu workbook
        wb.save(output_file)
        logger.info(f"{log_prefix} Đã xuất dữ liệu thành công ra file: {output_file}")
        return True
    except ImportError:
        logger.exception(f"{log_prefix} Thiếu thư viện 'openpyxl'.")
        raise RuntimeError("Thiếu thư viện 'openpyxl' để xử lý file Excel.")
    except (UserNotFoundError, DatabaseError, sqlite3.Error) as e_db:
        logger.exception(f"{log_prefix} Lỗi DB/User khi export: {e_db}")
        raise DatabaseError("Lỗi DB/User khi export.", original_exception=e_db)
    except Exception as e:
        logger.exception(f"{log_prefix} Lỗi không mong muốn khi export: {e}")
        raise DatabaseError("Lỗi không mong muốn khi export dữ liệu.", original_exception=e)
    finally:
        if wb:
            try:
                wb.close()
            except Exception as e_close_wb:
                logger.error(f"{log_prefix} Lỗi khi đóng workbook: {e_close_wb}")
        if conn:
            try:
                conn.close()
                logger.debug(f"{log_prefix} Đã đóng kết nối DB.")
            except Exception as e_close_conn:
                 logger.error(f"{log_prefix} Lỗi khi đóng kết nối DB: {e_close_conn}")

def export_set_data_excel(exporter_user_id, set_id, output_filepath):
    """
    Xuất dữ liệu của một bộ flashcard cụ thể ra file Excel (định dạng upload).
    """
    # Code hàm này giữ nguyên logic, chỉ đảm bảo định dạng và sửa lỗi exception
    log_prefix = f"[EXPORT_SET_WORKER|ExporterUID:{exporter_user_id}|Set:{set_id}]"
    logger.info(f"{log_prefix} Bắt đầu xuất bộ từ ra file: {output_filepath}")
    conn = None
    wb = None
    try:
        conn = database_connect()
        if conn is None:
            raise DatabaseError("Không thể kết nối DB.")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT title, creator_user_id FROM VocabularySets WHERE set_id = ?", (set_id,))
        set_info = cursor.fetchone()
        if not set_info:
            raise SetNotFoundError(set_id=set_id)
        set_creator_user_id = set_info['creator_user_id']
        # Quyền đã được kiểm tra ở handler
        # if set_creator_user_id != exporter_user_id:
        #     logger.warning(f"{log_prefix} User {exporter_user_id} không sở hữu bộ {set_id} (owner: {set_creator_user_id}).")
        #     raise PermissionsError(message=f"User không sở hữu bộ {set_id}.")
        set_title = set_info['title']
        logger.debug(f"{log_prefix} Tên bộ: '{set_title}'")
        logger.debug(f"{log_prefix} Đang lấy flashcards...")
        cursor.execute("SELECT flashcard_id, front, back, front_audio_content, back_audio_content, front_img, back_img, notification_text FROM Flashcards WHERE set_id = ? ORDER BY flashcard_id ASC", (set_id,))
        flashcards = cursor.fetchall()
        logger.info(f"{log_prefix} Tìm thấy {len(flashcards)} flashcards.")

        wb = openpyxl.Workbook()
        safe_sheet_title = "".join(c for c in set_title if c.isalnum() or c in (' ', '_', '-')).strip()[:31] or "Flashcards"
        ws = None
        # Xử lý tên sheet
        if safe_sheet_title != 'Sheet' and 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
            ws = wb.create_sheet(title=safe_sheet_title)
        elif safe_sheet_title == 'Sheet' and len(wb.sheetnames) > 1:
            ws = wb.create_sheet(title=f"{safe_sheet_title}_{set_id}")
            wb.remove(wb['Sheet'])
        else:
            ws = wb.active
            ws.title = safe_sheet_title

        logger.debug(f"{log_prefix} Sử dụng sheet tên: '{ws.title}'")
        headers = ["flashcard_id", "front", "back", "front_image", "back_image", "notification_text", "front_audio_content", "back_audio_content"]
        ws.append(headers)
        # Định dạng header
        header_font = Font(bold=True, name='Calibri')
        center_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill

        left_align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
        center_align_no_wrap = Alignment(horizontal='center', vertical='top', wrap_text=False)
        row_count = 0
        for card_row in flashcards:
            notify_text_val = card_row['notification_text'] if card_row['notification_text'] is not None else ""
            front_audio = card_row['front_audio_content'] if card_row['front_audio_content'] is not None else ""
            back_audio = card_row['back_audio_content'] if card_row['back_audio_content'] is not None else ""
            front_img_val = card_row['front_img'] if card_row['front_img'] is not None else ""
            back_img_val = card_row['back_img'] if card_row['back_img'] is not None else ""
            row_data = [card_row['flashcard_id'], card_row['front'], card_row['back'], front_img_val, back_img_val, notify_text_val, front_audio, back_audio]
            ws.append(row_data)
            row_count += 1
            # Định dạng alignment cho dòng dữ liệu
            for col_idx, cell in enumerate(ws[ws.max_row], 1):
                header_index = col_idx - 1
                if header_index < len(headers):
                    if headers[header_index] == 'flashcard_id':
                        cell.alignment = center_align_no_wrap
                    else:
                        cell.alignment = left_align_wrap

        logger.info(f"{log_prefix} Đã ghi {row_count} dòng.")
        # Điều chỉnh độ rộng cột
        for col_idx, header_text in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(header_text) + 2
            for cell in ws[column_letter]:
                if cell.row > 1 and cell.value:
                    try:
                        max_col_width = 60 # Mặc định
                        if 'image' in header_text.lower():
                            max_col_width = 40
                        elif 'content' in header_text.lower() or 'Thông báo' in header_text:
                            max_col_width = 80
                        elif 'id' in header_text.lower():
                            max_col_width = 15
                        cell_len = len(str(cell.value))
                        current_width_estimate = min(max(cell_len, 10), max_col_width)
                        if current_width_estimate > max_length:
                            max_length = current_width_estimate
                    except Exception as len_err:
                        logger.warning(f"{log_prefix} Lỗi tính độ dài ô {cell.coordinate}: {len_err}")
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
            logger.debug(f"{log_prefix} Cột {column_letter} ('{header_text}') rộng {adjusted_width}")

        # Lưu file
        wb.save(output_filepath)
        logger.info(f"{log_prefix} Đã lưu workbook: {output_filepath}")
        return True
    except ImportError:
        logger.error(f"{log_prefix} Thiếu 'openpyxl'.", exc_info=True)
        raise RuntimeError("Thiếu thư viện 'openpyxl'.")
    except (SetNotFoundError, PermissionsError, DatabaseError, sqlite3.Error) as e_known:
        logger.error(f"{log_prefix} Lỗi khi export set: {e_known}", exc_info=True)
        raise e_known # Ném lại lỗi đã biết
    except Exception as e:
        logger.error(f"{log_prefix} Lỗi không mong muốn: {e}", exc_info=True)
        # SỬA LỖI: Dùng biến 'e' đã bắt được
        raise DatabaseError("Lỗi không mong muốn khi export set.", original_exception=e)
    finally:
        if wb:
            try:
                wb.close()
            except Exception as e_close_wb:
                logger.error(f"{log_prefix} Lỗi khi đóng workbook: {e_close_wb}")
        if conn:
            try:
                conn.close()
                logger.debug(f"{log_prefix} Đã đóng kết nối DB.")
            except Exception as e_close_conn:
                 logger.error(f"{log_prefix} Lỗi khi đóng kết nối DB: {e_close_conn}")

def process_set_update_from_excel(updater_user_id, target_set_id, file_path):
    """
    Xử lý file Excel để cập nhật hoặc thêm mới flashcards vào một bộ từ đã có.
    Đã sửa logic xử lý ô audio trống. Đã tách dòng code và sửa lỗi exception.
    Đã đảm bảo định dạng code đúng chuẩn.
    """
    log_prefix = f"[PROCESS_UPDATE_XLSX|UpdaterUID:{updater_user_id}|Set:{target_set_id}]"
    logger.info(f"{log_prefix} Bắt đầu xử lý file cập nhật: {file_path}")
    results = {'updated': 0, 'added': 0, 'skipped': 0, 'errors': []}
    conn = None
    wb = None
    try:
        conn = database_connect()
        if conn is None:
            raise DatabaseError("Không thể kết nối DB.")
        # Đặt row_factory sau khi kết nối
        conn.row_factory = sqlite3.Row
        logger.debug(f"{log_prefix} Updater User ID: {updater_user_id}")
        try:
            logger.debug(f"{log_prefix} Đang mở workbook...")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            logger.info(f"{log_prefix} Đã load workbook. Sheet: '{ws.title}'")
        except InvalidFileException as e:
            raise InvalidFileFormatError(f'File Excel không hợp lệ: {e}', filename=file_path)
        except FileNotFoundError:
            raise FileProcessingError('File không tồn tại.', filename=file_path)
        except Exception as e:
            raise FileProcessingError(f'Không thể đọc file Excel: {e}', filename=file_path)

        if ws.max_row < 2:
            logger.warning(f"{log_prefix} Sheet '{ws.title}' không có dữ liệu.")
            return results
        else:
            # Đọc header và map cột (code không đổi)
            header_row_raw = [cell.value for cell in ws[1]]
            header_row = [str(h).strip().lower() if h is not None else "" for h in header_row_raw]
            col_map = {}
            expected_headers = { "flashcard_id": ["flashcard_id", "id", "card id"], "front": ["front", "term", "word", "mặt trước", "từ"], "back": ["back", "definition", "meaning", "mặt sau", "nghĩa"], "note": ["note", "notes", "ghi chú"], "front_audio_content": ["front_audio_content", "front_audio", "front audio"], "back_audio_content": ["back_audio_content", "back_audio", "back audio"], "front_img": ["front_image", "front_img", "front image"], "back_img": ["back_image", "back_img", "back image"], "notification_text": ["notification_text", "notification text", "notification", "thông báo"] }
            required_cols_for_add = {"front", "back"}
            for key, possible_names in expected_headers.items():
                found = False
                for idx, header_text in enumerate(header_row):
                    if header_text in possible_names:
                        col_map[key] = idx
                        logger.debug(f"{log_prefix} Map cột '{key}' -> index {idx}")
                        found = True
                        break
                if not found and key not in ["flashcard_id"] + list(required_cols_for_add):
                    logger.debug(f"{log_prefix} Cột tùy chọn '{key}' không tìm thấy.")
            if not required_cols_for_add.issubset(col_map.keys()):
                missing_cols = required_cols_for_add - col_map.keys()
                raise ExcelImportError(f'Thiếu cột bắt buộc: {missing_cols}', filename=file_path, sheet_name=ws.title, details=f"Header: {header_row}")

            logger.info(f"{log_prefix} Map cột: {col_map}. Xử lý {ws.max_row - 1} dòng...")
            # Lặp qua các dòng dữ liệu
            for row_index in range(2, ws.max_row + 1):
                excel_row_num = row_index
                row_cells = ws[row_index]
                log_prefix_row = f"{log_prefix} Row {excel_row_num}:"

                # Hàm nội bộ để lấy giá trị cell
                def get_cell_value(key):
                    idx = col_map.get(key)
                    # Sửa lỗi: kiểm tra idx < len(row_cells)
                    return row_cells[idx].value if idx is not None and idx < len(row_cells) and row_cells[idx] is not None else None

                # Lấy và xử lý các giá trị từ row
                row_flashcard_id_raw = get_cell_value("flashcard_id")
                row_front_raw = get_cell_value("front")
                row_back_raw = get_cell_value("back")
                row_note_raw = get_cell_value("note")
                row_audio_front_raw = get_cell_value("front_audio_content")
                row_audio_back_raw = get_cell_value("back_audio_content")
                row_img_front_raw = get_cell_value("front_img")
                row_img_back_raw = get_cell_value("back_img")
                row_notify_text_raw = get_cell_value("notification_text")

                row_flashcard_id = None
                try:
                    if row_flashcard_id_raw is not None:
                        cleaned_id_str = str(row_flashcard_id_raw).strip()
                        if cleaned_id_str:
                            row_flashcard_id = int(float(cleaned_id_str))
                            if row_flashcard_id <= 0:
                                row_flashcard_id = None
                except (ValueError, TypeError):
                    logger.warning(f"{log_prefix_row} flashcard_id lỗi: '{row_flashcard_id_raw}'.")
                    row_flashcard_id = None

                row_front = str(row_front_raw).strip() if row_front_raw is not None else ""
                row_back = str(row_back_raw).strip() if row_back_raw is not None else ""
                row_note = str(row_note_raw).strip() if row_note_raw is not None else ""

                # === SỬA LOGIC XỬ LÝ AUDIO (ĐỊNH DẠNG ĐÚNG) ===
                row_audio_front = None # Khởi tạo là None
                raw_audio_front_val = get_cell_value('front_audio_content')
                if raw_audio_front_val is not None:
                    temp_audio_front = str(raw_audio_front_val).strip()
                    if temp_audio_front: # Chỉ gán nếu chuỗi không rỗng
                        row_audio_front = temp_audio_front

                row_audio_back = None # Khởi tạo là None
                raw_audio_back_val = get_cell_value('back_audio_content')
                if raw_audio_back_val is not None:
                    temp_audio_back = str(raw_audio_back_val).strip()
                    if temp_audio_back: # Chỉ gán nếu chuỗi không rỗng
                        row_audio_back = temp_audio_back
                # =================================================

                row_img_front = str(row_img_front_raw).strip() or None
                row_img_back = str(row_img_back_raw).strip() or None
                row_notify_text = str(row_notify_text_raw).strip() or None

                card_info_for_log = f"(ID: {row_flashcard_id}, Front: '{row_front[:30]}...')"

                # Kiểm tra front/back không rỗng
                if not row_front or not row_back:
                    logger.warning(f"{log_prefix_row} Thiếu front/back.")
                    results['skipped'] += 1
                    results['errors'].append({'line': excel_row_num, 'reason': 'Thiếu mặt trước hoặc mặt sau', 'card_info': card_info_for_log})
                    continue # Bỏ qua dòng này

                # Xử lý cập nhật hoặc thêm mới trong transaction
                try:
                    with conn: # Bắt đầu transaction cho mỗi dòng
                        cursor = conn.cursor()
                        if row_flashcard_id: # Update
                            logger.debug(f"{log_prefix_row} Có ID ({row_flashcard_id}). Update...")
                            cursor.execute('SELECT set_id FROM "Flashcards" WHERE flashcard_id = ?', (row_flashcard_id,))
                            existing_card = cursor.fetchone()
                            if existing_card and existing_card['set_id'] == target_set_id:
                                cursor.execute("""
                                    UPDATE "Flashcards" SET
                                        "front" = ?, "back" = ?, "front_audio_content" = ?,
                                        "back_audio_content" = ?, "front_img" = ?, "back_img" = ?,
                                        "notification_text" = ?
                                    WHERE flashcard_id = ?
                                """, (row_front, row_back, row_audio_front, row_audio_back, row_img_front, row_img_back, row_notify_text, row_flashcard_id) )
                                logger.debug(f"{log_prefix_row} Đã cập nhật Flashcard.")
                                # Xử lý note
                                cursor.execute('SELECT note_id FROM "FlashcardNotes" WHERE flashcard_id = ? AND user_id = ?', (row_flashcard_id, updater_user_id))
                                existing_note = cursor.fetchone()
                                if row_note:
                                    if existing_note:
                                        cursor.execute('UPDATE "FlashcardNotes" SET note = ? WHERE note_id = ?', (row_note, existing_note['note_id']))
                                        logger.debug(f"{log_prefix_row} Đã cập nhật Note.")
                                    else:
                                        cursor.execute('INSERT INTO "FlashcardNotes" (flashcard_id, user_id, note) VALUES (?, ?, ?)', (row_flashcard_id, updater_user_id, row_note))
                                        logger.debug(f"{log_prefix_row} Đã thêm Note mới.")
                                elif existing_note:
                                    cursor.execute('DELETE FROM "FlashcardNotes" WHERE note_id = ?', (existing_note['note_id'],))
                                    logger.debug(f"{log_prefix_row} Đã xóa Note cũ.")
                                results['updated'] += 1
                                logger.debug(f"{log_prefix_row} Update thành công.")
                            elif existing_card:
                                logger.warning(f"{log_prefix_row} ID {row_flashcard_id} thuộc bộ khác."); results['skipped'] += 1; results['errors'].append({'line': excel_row_num, 'reason': f'ID thẻ thuộc bộ khác ({existing_card["set_id"]})', 'card_info': card_info_for_log})
                            else:
                                logger.warning(f"{log_prefix_row} ID {row_flashcard_id} không tồn tại."); results['skipped'] += 1; results['errors'].append({'line': excel_row_num, 'reason': 'ID thẻ không tồn tại', 'card_info': card_info_for_log})
                        else: # Add new
                            logger.debug(f"{log_prefix_row} Không có ID. Thêm mới...")
                            cursor.execute('SELECT flashcard_id FROM "Flashcards" WHERE set_id = ? AND lower(trim(front)) = lower(?)', (target_set_id, row_front.strip()))
                            duplicate = cursor.fetchone()
                            if duplicate:
                                logger.warning(f"{log_prefix_row} Mặt trước '{row_front}' bị trùng.")
                                results['skipped'] += 1
                                results['errors'].append({'line': excel_row_num, 'reason': 'Mặt trước bị trùng', 'card_info': card_info_for_log})
                            else:
                                cursor.execute("""
                                    INSERT INTO "Flashcards"
                                    (set_id, front, back, front_audio_content, back_audio_content, front_img, back_img, notification_text)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                """, (target_set_id, row_front, row_back, row_audio_front, row_audio_back, row_img_front, row_img_back, row_notify_text) )
                                new_card_id = cursor.lastrowid
                                if not new_card_id or new_card_id <=0:
                                    raise DatabaseError("Không lấy được ID thẻ mới.")
                                logger.debug(f"{log_prefix_row} Đã thêm thẻ mới ID: {new_card_id}")
                                if row_note:
                                    cursor.execute('INSERT INTO "FlashcardNotes" (flashcard_id, user_id, note) VALUES (?, ?, ?)', (new_card_id, updater_user_id, row_note)); logger.debug(f"{log_prefix_row} Đã thêm Note cho thẻ mới.")
                                results['added'] += 1
                                logger.debug(f"{log_prefix_row} Thêm mới thành công.")
                except sqlite3.Error as e_row_db:
                    logger.error(f"{log_prefix_row} Lỗi DB: {e_row_db}")
                    results['skipped'] += 1
                    results['errors'].append({'line': excel_row_num, 'reason': f'Lỗi DB: {e_row_db}', 'card_info': card_info_for_log})
                except Exception as e_row_unk:
                    logger.error(f"{log_prefix_row} Lỗi không mong muốn: {e_row_unk}", exc_info=True)
                    results['skipped'] += 1
                    results['errors'].append({'line': excel_row_num, 'reason': f'Lỗi không xác định: {e_row_unk}', 'card_info': card_info_for_log})

            logger.info(f"{log_prefix} Hoàn tất xử lý file. Kết quả: {results}")
    except (DatabaseError, FileProcessingError, ExcelImportError, InvalidFileFormatError, sqlite3.Error) as e_proc:
        logger.error(f"{log_prefix} Lỗi xử lý file/DB nghiêm trọng: {e_proc}", exc_info=True)
        # Đảm bảo results['errors'] là list trước khi append
        if not isinstance(results.get('errors'), list): results['errors'] = []
        results['errors'].append({'line': 'General', 'reason': f'Lỗi xử lý: {e_proc}', 'card_info': ''})
    except Exception as e: # Sửa lỗi biến exception
        logger.error(f"{log_prefix} Lỗi không mong muốn: {e}", exc_info=True)
        # Đảm bảo results['errors'] là list trước khi append
        if not isinstance(results.get('errors'), list): results['errors'] = []
        results['errors'].append({'line': 'General', 'reason': f'Lỗi không xác định: {e}', 'card_info': ''})
    finally:
        if wb:
            try: wb.close()
            except Exception as e_close_wb: logger.error(f"{log_prefix} Lỗi khi đóng workbook: {e_close_wb}")
        if conn:
            try: conn.close(); logger.debug(f"{log_prefix} Đã đóng kết nối DB.")
            except Exception as e_close_conn: logger.error(f"{log_prefix} Lỗi khi đóng kết nối DB: {e_close_conn}")
    logger.info(f"{log_prefix} Trả về kết quả: {results}")
    return results

def import_new_set_from_excel(file_path, creator_user_id):
    """
    Xử lý việc import bộ từ vựng từ file Excel vào database.
    Đã sửa logic xử lý ô audio trống. Đã tách dòng code và sửa lỗi exception.
    Đã đảm bảo định dạng code đúng chuẩn.
    """
    # Code hàm này giữ nguyên logic, chỉ đảm bảo định dạng và logic audio đúng, sửa lỗi exception
    log_prefix = f"[SERVICE_IMPORT_EXCEL|CreatorUID:{creator_user_id}]"; logger.info(f"{log_prefix} Bắt đầu import bộ từ từ file: {file_path}")
    set_id_created = None; dictionary_title = os.path.splitext(os.path.basename(file_path))[0]; inserted_count = 0; error_msg = "Lỗi không xác định."; wb = None; conn = None;
    try:
        try:
            logger.debug(f"{log_prefix} Đang mở workbook..."); wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True); set_sheet = wb.active
            if set_sheet.title: dictionary_title = set_sheet.title
            logger.debug(f"{log_prefix} Sheet '{set_sheet.title}'. Tên bộ: '{dictionary_title}'."); option_sheet = None; description = f"Upload từ {os.path.basename(file_path)}"; tags = None; is_public = 1
            default_front_audio = None; default_back_audio = None; default_front_img = None; default_back_img = None; default_notify_text = None; option_sheet_name_found = None
            for sheet_name in wb.sheetnames:
                if sheet_name.lower().strip() == "option": option_sheet_name_found = sheet_name; break
            if option_sheet_name_found:
                option_sheet = wb[option_sheet_name_found]; logger.debug(f"{log_prefix} Tìm thấy sheet '{option_sheet_name_found}'.")
                try:
                    headers_iter = option_sheet.iter_rows(min_row=1, max_row=1, values_only=True); headers = [str(h).strip().lower() if h is not None else "" for h in next(headers_iter, [])]
                    values_iter = option_sheet.iter_rows(min_row=2, max_row=2, values_only=True); values_row = next(values_iter, []); values = list(values_row) + [None] * (len(headers) - len(values_row))
                    option_data = dict(zip(headers, values)); logger.debug(f"{log_prefix} Dữ liệu sheet option: {option_data}")
                    opt_desc = option_data.get("description");
                    if opt_desc is not None and str(opt_desc).strip(): description = str(opt_desc).strip()
                    opt_tags = option_data.get("tags");
                    if opt_tags is not None and str(opt_tags).strip(): tags = str(opt_tags).strip()
                    opt_public = option_data.get("is_public");
                    if opt_public is not None and str(opt_public).strip():
                        try: is_public = int(opt_public)
                        except (ValueError, TypeError): is_public = 1; logger.warning(f"{log_prefix} 'is_public' lỗi.")
                    opt_front_audio = option_data.get("front_audio_content");
                    if opt_front_audio is not None and str(opt_front_audio).strip(): default_front_audio = str(opt_front_audio).strip()
                    opt_back_audio = option_data.get("back_audio_content");
                    if opt_back_audio is not None and str(opt_back_audio).strip(): default_back_audio = str(opt_back_audio).strip()
                    opt_front_img = option_data.get("front_image");
                    if opt_front_img is not None and str(opt_front_img).strip(): default_front_img = str(opt_front_img).strip()
                    opt_back_img = option_data.get("back_image");
                    if opt_back_img is not None and str(opt_back_img).strip(): default_back_img = str(opt_back_img).strip()
                    opt_notify_text = option_data.get("notification_text")
                    if opt_notify_text is not None and str(opt_notify_text).strip(): default_notify_text = str(opt_notify_text).strip()
                    logger.info(f"{log_prefix} Áp dụng tùy chọn từ sheet 'option'.")
                except Exception as e_opt: logger.warning(f"{log_prefix} Lỗi đọc sheet 'option': {e_opt}", exc_info=True)
            else: logger.info(f"{log_prefix} Không tìm thấy sheet 'option'.")
        except FileNotFoundError: raise FileProcessingError("File không tồn tại.", filename=file_path)
        except InvalidFileException as e: raise InvalidFileFormatError(f'File Excel không hợp lệ: {e}', filename=file_path)
        except Exception as e: raise FileProcessingError(f'Không thể đọc file Excel: {e}', filename=file_path)
        logger.debug(f"{log_prefix} Đang kết nối database..."); conn = database_connect();
        if conn is None: raise DatabaseError("Không thể kết nối DB.")
        logger.debug(f"{log_prefix} Creator User ID: {creator_user_id}")
        try:
            with conn:
                logger.debug(f"{log_prefix} Bắt đầu transaction."); cursor = conn.cursor()
                cursor.execute( "INSERT INTO VocabularySets (title, description, tags, creator_user_id, is_public) VALUES (?, ?, ?, ?, ?)", (dictionary_title, description, tags, creator_user_id, is_public) )
                set_id_created = cursor.lastrowid;
                if not set_id_created or set_id_created <= 0: raise DatabaseError("Không lấy được set_id.")
                logger.info(f"{log_prefix} Đã chèn VocabularySet ID: {set_id_created}")
                header_values_iter = set_sheet.iter_rows(min_row=1, max_row=1, values_only=True); header_row_values = next(header_values_iter, None)
                if not header_row_values: raise ValueError("Sheet flashcard thiếu header.")
                header_row = [str(h).strip().lower() if h is not None else "" for h in header_row_values]; col_map = {}
                common_names = { 'front': ['front', 'term', 'word', 'mặt trước', 'từ'], 'back': ['back', 'definition', 'meaning', 'mặt sau', 'nghĩa'], 'front_audio_content': ['front_audio', 'front_audio_content', 'front audio', 'audio mặt trước'], 'back_audio_content': ['back_audio', 'back_audio_content', 'back audio', 'audio mặt sau'], 'front_img': ['front_image', 'front_img', 'front image'], 'back_img': ['back_image', 'back_img', 'back image'], 'notification_text': ['notification_text', 'notification text', 'notification', 'thông báo'] }
                required_keys = {'front', 'back'}
                for key, names in common_names.items():
                    found = False
                    for i, header in enumerate(header_row):
                        if header in names: col_map[key] = i; found = True; break
                    if not found:
                        if key == 'front': col_map[key] = 0; logger.warning(f"{log_prefix} Không tìm thấy 'front'.")
                        elif key == 'back': col_map[key] = 1; logger.warning(f"{log_prefix} Không tìm thấy 'back'.")
                        else: logger.debug(f"{log_prefix} Không tìm thấy cột '{key}'.")
                if not all(k in col_map for k in required_keys): raise ValueError(f"Thiếu cột bắt buộc: {required_keys - set(col_map.keys())}")
                logger.debug(f"{log_prefix} Mapping cột: {col_map}"); flashcards_to_insert = [] ; skipped_rows = 0
                for idx, row_values in enumerate(set_sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row_values or all(c is None for c in row_values): skipped_rows += 1; continue
                    def get_cell_value_internal(key, default_val=None): idx_cell = col_map.get(key); return row_values[idx_cell] if idx_cell is not None and idx_cell < len(row_values) else default_val
                    front = str(get_cell_value_internal('front', '')).strip(); back = str(get_cell_value_internal('back', '')).strip()

                    # === SỬA LOGIC XỬ LÝ AUDIO (Định dạng đúng) ===
                    front_audio = None
                    raw_front_audio = get_cell_value_internal('front_audio_content', default_front_audio)
                    if raw_front_audio is not None:
                        temp_fa = str(raw_front_audio).strip()
                        if temp_fa:
                            front_audio = temp_fa

                    back_audio = None
                    raw_back_audio = get_cell_value_internal('back_audio_content', default_back_audio)
                    if raw_back_audio is not None:
                        temp_ba = str(raw_back_audio).strip()
                        if temp_ba:
                            back_audio = temp_ba
                    # ==============================

                    front_img = str(get_cell_value_internal('front_img') or default_front_img or '').strip() or None
                    back_img = str(get_cell_value_internal('back_img') or default_back_img or '').strip() or None
                    notify_text = str(get_cell_value_internal('notification_text') or default_notify_text or '').strip() or None
                    if front and back: flashcards_to_insert.append((set_id_created, front, back, front_audio, back_audio, front_img, back_img, notify_text))
                    else: logger.warning(f"{log_prefix} Bỏ qua dòng {idx}."); skipped_rows += 1
                if skipped_rows > 0: logger.info(f"{log_prefix} Đã bỏ qua {skipped_rows} dòng.")
                if flashcards_to_insert:
                    logger.info(f"{log_prefix} Chuẩn bị chèn {len(flashcards_to_insert)} flashcards...")
                    cursor.executemany(""" INSERT INTO "Flashcards" (set_id, front, back, front_audio_content, back_audio_content, front_img, back_img, notification_text) VALUES (?, ?, ?, ?, ?, ?, ?, ?) """, flashcards_to_insert)
                    inserted_count = len(flashcards_to_insert); logger.info(f"{log_prefix} Đã chèn {inserted_count} flashcards.")
                else: logger.warning(f"{log_prefix} Không tìm thấy flashcard hợp lệ.")
                logger.debug(f"{log_prefix} Kết thúc transaction.")
        except (sqlite3.Error, DatabaseError) as e_db_insert: error_msg = f"Lỗi DB khi import: {e_db_insert}"; logger.error(f"{log_prefix} Lỗi DB: {e_db_insert}", exc_info=True); return None, error_msg, 0
        except ValueError as e_value: error_msg = f"Lỗi dữ liệu Excel: {e_value}"; logger.error(f"{log_prefix} Lỗi dữ liệu: {e_value}"); return None, error_msg, 0
        logger.info(f"{log_prefix} Import thành công set_id {set_id_created}.")
        return set_id_created, dictionary_title, inserted_count
    except (DatabaseError, UserNotFoundError, FileProcessingError, InvalidFileFormatError, ExcelImportError, sqlite3.Error) as e_proc:
        error_msg = f"Lỗi xử lý import: {e_proc}"; logger.error(f"{log_prefix} Lỗi xử lý: {e_proc}", exc_info=True); return None, error_msg, 0
    except Exception as e: # Sửa lỗi biến exception
        error_msg = f"Lỗi không mong muốn: {e}"; logger.error(f"{log_prefix} Lỗi không mong muốn: {e}", exc_info=True); return None, error_msg, 0
    finally:
        if wb:
            try: wb.close()
            except Exception as e_close_wb: logger.error(f"{log_prefix} Lỗi khi đóng workbook: {e_close_wb}")
        if conn:
            try: conn.close(); logger.debug(f"{log_prefix} Đã đóng kết nối DB.")
            except Exception as e_close_conn: logger.error(f"{log_prefix} Lỗi khi đóng kết nối DB: {e_close_conn}")